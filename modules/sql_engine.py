import csv
import os
import re
import sqlite3
import threading
from typing import List, Dict, Any, Tuple, Optional

try:
    import openpyxl  # optional for .xlsx
except Exception:
    openpyxl = None

class SimpleSQLite:
    def __init__(self, table_name: str):
        self.table_name = table_name
        self.conn = sqlite3.connect(":memory:", check_same_thread=False)
        self.lock = threading.Lock()
        self.conn.row_factory = sqlite3.Row

    def _infer_type(self, s: str) -> str:
        if s is None:
            return "TEXT"
        s = str(s)
        # int?
        if re.fullmatch(r"[+-]?\d+", s):
            return "INTEGER"
        # float?
        if re.fullmatch(r"[+-]?(\d*\.\d+|\d+\.\d*)([eE][+-]?\d+)?", s):
            return "REAL"
        # iso date?
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
            return "TEXT"
        return "TEXT"

    def load_rows(self, rows: List[Dict[str, Any]]):
        if not rows:
            raise ValueError("No rows to load")
        # infer types from first non-empty row
        cols = list(rows[0].keys())
        types = []
        for c in cols:
            v = next((r.get(c) for r in rows if r.get(c) not in (None, "")), None)
            types.append(self._infer_type(v) if v is not None else "TEXT")

        col_defs = ", ".join(f'"{c}" {t}' for c, t in zip(cols, types))
        with self.lock:
            self.conn.execute(f'DROP TABLE IF EXISTS "{self.table_name}"')
            self.conn.execute(f'CREATE TABLE "{self.table_name}" ({col_defs})')

            placeholders = ", ".join("?" for _ in cols)
            insert_sql = f'INSERT INTO "{self.table_name}" ({", ".join([f'"{c}"' for c in cols])}) VALUES ({placeholders})'
            for r in rows:
                values = [r.get(c, None) for c in cols]
                self.conn.execute(insert_sql, values)
            self.conn.commit()

    def execute_safe_select(self, sql: str) -> Tuple[List[str], List[Dict[str, Any]]]:
        s = sql.strip().lower()
        if not (s.startswith("select") or s.startswith("with")):
            raise ValueError("Only SELECT/WITH queries are allowed in demo mode.")
        with self.lock:
            cur = self.conn.execute(sql)
            col_names = [d[0] for d in cur.description]
            data = [dict(row) for row in cur.fetchall()]
        return col_names, data

    def schema_text(self, sample_rows: int = 3) -> str:
        # build CREATE TABLE-ish schema description
        with self.lock:
            info = self.conn.execute(f'PRAGMA table_info("{self.table_name}")').fetchall()
            cols = [(row[1], row[2]) for row in info]  # name, type
            sample = self.conn.execute(f'SELECT * FROM "{self.table_name}" LIMIT {sample_rows}').fetchall()
        lines = [f'CREATE TABLE "{self.table_name}" (']
        for name, typ in cols:
            lines.append(f'  "{name}" {typ},')
        lines[-1] = lines[-1].rstrip(",")
        lines.append(");")
        if sample:
            lines.append("\n-- Sample rows:")
            # show a couple of rows
            for r in sample:
                as_dict = {k: r[k] for k in r.keys()}
                lines.append(str(as_dict))
        return "\n".join(lines)

def load_csv(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader)

def load_xlsx(path: str) -> List[Dict[str, Any]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed; cannot read .xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    out = []
    for r in rows[1:]:
        out.append({h: v for h, v in zip(headers, r)})
    return out

def build_engine_from_file(path: str, table_name: str) -> SimpleSQLite:
    _, ext = os.path.splitext(path.lower())
    if ext == ".csv":
        rows = load_csv(path)
    elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        rows = load_xlsx(path)
    else:
        raise ValueError("Unsupported file type; use .csv or .xlsx")
    engine = SimpleSQLite(table_name=table_name)
    engine.load_rows(rows)
    return engine
